#DEV 18/NOV/2022

import openpyxl as opxl

def eliminarFilas(dir):

    wb = opxl.load_workbook(dir)
    defecto = "Evaluation Warning" #delete default sheet
    if defecto in wb.sheetnames:
        wb.remove(wb[defecto])
        wb.save(dir)

    wb = opxl.load_workbook(dir)
    sheet = wb.active
    for i in range(10):

        if(sheet["A1"].value == None):
            sheet.delete_rows(1)
        else:
            break
        
    wb.save(dir)
